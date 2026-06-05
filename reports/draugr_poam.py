"""
draugr_poam.py — DoD/RMF Plan of Action & Milestones (POA&M) XLSX exporter.

Generates a formatted Excel workbook following the standard DoD POA&M
column structure, pre-populated from Draugr scan results.

Column mapping follows DISA STIG/RMF POA&M template conventions.
"""
import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ------------------------------------------------------------------
# Colour palette (matches Draugr dark theme adapted for XLSX)
# ------------------------------------------------------------------
_CLR_HEADER_BG   = "1F0D0D"   # deep dark red
_CLR_HEADER_FG   = "E4CCC8"   # warm off-white
_CLR_CRITICAL    = "CB322C"   # Draugr crimson
_CLR_HIGH        = "D4722A"   # burnt orange
_CLR_MEDIUM      = "C4935A"   # amber
_CLR_LOW         = "7A9E6E"   # muted green
_CLR_INFO        = "7A8FA3"   # steel blue
_CLR_KEV         = "FF4444"   # bright red for KEV flags
_CLR_ROW_ALT     = "1A0808"   # very subtle row striping
_CLR_ROW_NORMAL  = "170909"   # background
_CLR_BORDER      = "551E1E"   # dark crimson border


def _to_float(v: Any) -> float:
    try:
        return float(str(v).replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _sev_fill(sev: str) -> Optional["PatternFill"]:
    if not HAS_OPENPYXL:
        return None
    colours = {
        "CRITICAL": _CLR_CRITICAL,
        "HIGH":     _CLR_HIGH,
        "MEDIUM":   _CLR_MEDIUM,
        "LOW":      _CLR_LOW,
    }
    c = colours.get(sev.upper())
    if c:
        return PatternFill(fill_type="solid", fgColor=c)
    return None


def _scheduled_completion(r: Dict[str, Any]) -> str:
    """
    Return a scheduled completion date based on severity/KEV status.
    Tier 1 (KEV/critical+exploit):  +3 days
    Tier 2 (critical or exploit):   +7 days
    Tier 3 (high):                  +30 days
    Tier 4 (medium/low):            +90 days
    """
    today = datetime.date.today()
    kev   = str(r.get("Known Exploited Vulnerability","")).upper() == "YES"
    expl  = str(r.get("Public Exploit","")).upper() == "YES"
    sev   = str(r.get("CVSS Severity","") or "").upper()
    rs    = _to_float(r.get("Risk Score", 0))

    if kev or (sev == "CRITICAL" and expl):
        delta = 3
    elif sev == "CRITICAL" or expl:
        delta = 7
    elif sev == "HIGH":
        delta = 30
    else:
        delta = 90

    return (today + datetime.timedelta(days=delta)).isoformat()


def _poam_id(idx: int) -> str:
    return f"DRAUGR-{idx:04d}"


def _control_family(nist_str: str) -> str:
    """Extract the primary NIST control family from a semicolon-delimited string."""
    if not nist_str:
        return ""
    first = nist_str.split(";")[0].strip().split()[0]
    family_map = {
        "AC": "Access Control",  "AT": "Awareness & Training",
        "AU": "Audit & Accountability", "CA": "Assessment, Auth & Monitoring",
        "CM": "Config Management", "CP": "Contingency Planning",
        "IA": "Identification & Auth", "IR": "Incident Response",
        "MA": "Maintenance", "MP": "Media Protection",
        "PE": "Physical Protection", "PL": "Planning",
        "PM": "Program Management", "PS": "Personnel Security",
        "PT": "PII Processing", "RA": "Risk Assessment",
        "SA": "System & Services Acquisition", "SC": "System & Comms Protection",
        "SI": "System & Info Integrity", "SR": "Supply Chain",
    }
    fam = first.split("-")[0].upper() if "-" in first else first.upper()
    return f"{first} — {family_map.get(fam, 'See NIST SP 800-53')}"


# ------------------------------------------------------------------
# Main export function
# ------------------------------------------------------------------
def export_poam(
    rows: List[Dict[str, Any]],
    output_path: str,
    system_name: str = "System Under Assessment",
    prepared_by: str = "Draugr Threat Intelligence System",
    classification: str = "UNCLASSIFIED",
) -> int:
    """
    Export scan rows to a DoD/RMF-format POA&M XLSX workbook.
    Returns the number of POA&M entries written.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for POA&M export. "
            "Install it with: pip install openpyxl"
        )

    today = datetime.date.today().isoformat()

    # Filter to only actionable findings (exclude info/resolved)
    actionable = [
        r for r in rows
        if str(r.get("CVSS Severity","") or "").upper() in
           ("CRITICAL","HIGH","MEDIUM","LOW")
    ]
    # Sort by risk score descending
    actionable.sort(key=lambda r: _to_float(r.get("Risk Score", 0)), reverse=True)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Cover / Summary
    # ------------------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Summary"

    header_font  = Font(name="Calibri", bold=True, color=_CLR_HEADER_FG, size=12)
    header_fill  = PatternFill(fill_type="solid", fgColor=_CLR_HEADER_BG)
    value_font   = Font(name="Calibri", size=11, color="E4CCC8")
    title_font   = Font(name="Calibri", bold=True, size=16, color=_CLR_CRITICAL)
    sub_font     = Font(name="Calibri", size=10, color="7A5C5A", italic=True)

    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 48

    ws_cover["A1"] = "PLAN OF ACTION & MILESTONES (POA&M)"
    ws_cover["A1"].font = title_font
    ws_cover["A1"].fill = header_fill
    ws_cover.merge_cells("A1:B1")

    ws_cover["A2"] = f"Classification: {classification}"
    ws_cover["A2"].font = sub_font
    ws_cover.merge_cells("A2:B2")

    summary_data = [
        ("System / Asset Name",    system_name),
        ("Prepared By",            prepared_by),
        ("Date Prepared",          today),
        ("Total Findings",         str(len(actionable))),
        ("Critical",               str(sum(1 for r in actionable if str(r.get("CVSS Severity","")).upper()=="CRITICAL"))),
        ("High",                   str(sum(1 for r in actionable if str(r.get("CVSS Severity","")).upper()=="HIGH"))),
        ("Medium",                 str(sum(1 for r in actionable if str(r.get("CVSS Severity","")).upper()=="MEDIUM"))),
        ("Low",                    str(sum(1 for r in actionable if str(r.get("CVSS Severity","")).upper()=="LOW"))),
        ("CISA KEV Listed",        str(sum(1 for r in actionable if str(r.get("Known Exploited Vulnerability","")).upper()=="YES"))),
        ("Public Exploit Code",    str(sum(1 for r in actionable if str(r.get("Public Exploit","")).upper()=="YES"))),
        ("Tier 1 (24–72 hr)",      str(sum(1 for r in actionable if _scheduled_completion(r) <= (datetime.date.today()+datetime.timedelta(days=3)).isoformat()))),
        ("Tier 2 (7 days)",        str(sum(1 for r in actionable if (datetime.date.today()+datetime.timedelta(days=3)).isoformat() < _scheduled_completion(r) <= (datetime.date.today()+datetime.timedelta(days=7)).isoformat()))),
        ("Tier 3 (30 days)",       str(sum(1 for r in actionable if (datetime.date.today()+datetime.timedelta(days=7)).isoformat() < _scheduled_completion(r) <= (datetime.date.today()+datetime.timedelta(days=30)).isoformat()))),
        ("Tier 4 (90 days)",       str(sum(1 for r in actionable if _scheduled_completion(r) > (datetime.date.today()+datetime.timedelta(days=30)).isoformat()))),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=4):
        cell_l = ws_cover.cell(row=row_idx, column=1, value=label)
        cell_v = ws_cover.cell(row=row_idx, column=2, value=value)
        cell_l.font = header_font
        cell_l.fill = header_fill
        cell_v.font = value_font
        cell_v.fill = PatternFill(fill_type="solid", fgColor=_CLR_ROW_NORMAL)

    # ------------------------------------------------------------------
    # Sheet 2: POA&M Findings
    # ------------------------------------------------------------------
    ws = wb.create_sheet("POA&M Findings")

    col_defs = [
        # (header, width, field_or_callable)
        ("POA&M ID",               14, lambda i, r: _poam_id(i)),
        ("Weakness / Finding",     45, lambda i, r: str(r.get("Description","") or "")[:300]),
        ("CVE ID",                 16, lambda i, r: r.get("CVE ID","")),
        ("Severity",               12, lambda i, r: r.get("CVSS Severity","")),
        ("Risk Score",             12, lambda i, r: r.get("Risk Score","")),
        ("CVSS Base Score",        14, lambda i, r: r.get("CVSS Base Score","")),
        ("EPSS",                   10, lambda i, r: r.get("EPSS Score","")),
        ("CISA KEV",               10, lambda i, r: r.get("Known Exploited Vulnerability","No")),
        ("Public Exploit",         14, lambda i, r: r.get("Public Exploit","No")),
        ("Software",               30, lambda i, r: f"{r.get('Software Name','')} {r.get('Software Version','')}".strip()),
        ("Publisher",              28, lambda i, r: r.get("Publisher","")),
        ("Install Date",           14, lambda i, r: r.get("Install Date","")),
        ("Patch Age (Days)",       14, lambda i, r: r.get("Patch Age (Days)","")),
        ("Version Confirmed",      16, lambda i, r: r.get("Version Confirmed","")),
        ("NIST Control",           30, lambda i, r: _control_family(str(r.get("NIST 800-53 Controls","") or ""))),
        ("ATT&CK Technique",       30, lambda i, r: str(r.get("ATT&CK Techniques","") or "")[:100]),
        ("CWE",                    20, lambda i, r: str(r.get("CWE","") or "")[:60]),
        ("Vendor Advisory",        45, lambda i, r: r.get("Vendor Advisory URL","") or r.get("NVD URL","")),
        ("Responsible Party",      24, lambda i, r: "IT Security / ISSO"),
        ("Resources Required",     24, lambda i, r: "Patch management cycle"),
        ("Scheduled Completion",   20, lambda i, r: _scheduled_completion(r)),
        ("Actual Completion",      18, lambda i, r: ""),
        ("Status",                 16, lambda i, r: "Open"),
        ("Milestones / Notes",     40, lambda i, r: _milestone_note(r)),
    ]

    # Header row
    header_row_fill = PatternFill(fill_type="solid", fgColor=_CLR_HEADER_BG)
    thin_border = Border(
        bottom=Side(style="thin", color=_CLR_BORDER),
        right=Side(style="thin", color=_CLR_BORDER),
    )

    for col_idx, (col_header, col_width, _) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_header)
        cell.font  = Font(name="Calibri", bold=True, color=_CLR_HEADER_FG, size=10)
        cell.fill  = header_row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, r in enumerate(actionable, start=2):
        poam_index = row_idx - 1
        alt = (row_idx % 2 == 0)
        row_bg = _CLR_ROW_ALT if alt else _CLR_ROW_NORMAL
        row_fill = PatternFill(fill_type="solid", fgColor=row_bg)
        sev = str(r.get("CVSS Severity","") or "").upper()
        kev = str(r.get("Known Exploited Vulnerability","")).upper() == "YES"

        for col_idx, (_, _, getter) in enumerate(col_defs, start=1):
            val  = getter(poam_index, r)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10, color="E4CCC8")
            cell.fill      = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = thin_border

            col_header = col_defs[col_idx - 1][0]

            # Severity cell coloring
            if col_header == "Severity" and sev:
                sf = _sev_fill(sev)
                if sf:
                    cell.fill = sf
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

            # KEV highlight
            if col_header == "CISA KEV" and kev:
                cell.fill = PatternFill(fill_type="solid", fgColor=_CLR_KEV)
                cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

            # Hyperlink for advisory/NVD URL
            if col_header == "Vendor Advisory" and val and val.startswith("http"):
                cell.hyperlink = val
                cell.value = val
                cell.font = Font(name="Calibri", size=10, color="7A8FA3", underline="single")

            # Hyperlink for CVE ID
            if col_header == "CVE ID" and val and val.startswith("CVE-"):
                nvd_url = f"https://nvd.nist.gov/vuln/detail/{val}"
                cell.hyperlink = nvd_url
                cell.font = Font(name="Calibri", size=10, color="7A8FA3", underline="single")

        ws.row_dimensions[row_idx].height = 48

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}1"

    # ------------------------------------------------------------------
    # Sheet 3: False Positive Suppression Log
    # ------------------------------------------------------------------
    ws_fp = wb.create_sheet("False Positives")
    fp_headers = ["CVE ID", "Software Name", "Reason", "Date Added", "Accepted By"]
    for col_idx, h in enumerate(fp_headers, start=1):
        cell = ws_fp.cell(row=1, column=col_idx, value=h)
        cell.font  = Font(name="Calibri", bold=True, color=_CLR_HEADER_FG, size=10)
        cell.fill  = header_row_fill
    ws_fp.column_dimensions["A"].width = 18
    ws_fp.column_dimensions["B"].width = 35
    ws_fp.column_dimensions["C"].width = 45
    ws_fp.column_dimensions["D"].width = 14
    ws_fp.column_dimensions["E"].width = 20
    # Data will be filled in by the caller or left blank for manual completion

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(actionable)


def _milestone_note(r: Dict[str, Any]) -> str:
    """Generate a standard milestone note from the finding data."""
    parts = []
    kev  = str(r.get("Known Exploited Vulnerability","")).upper() == "YES"
    expl = str(r.get("Public Exploit","")).upper() == "YES"
    sev  = str(r.get("CVSS Severity","") or "").upper()
    pa   = str(r.get("Patch Age (Days)","") or "")
    pub  = str(r.get("Publisher","") or "")

    if kev:
        parts.append("CISA KEV-listed — emergency patching required per BOD 22-01.")
    if expl:
        parts.append("Public exploit code exists — prioritize above standard patch SLA.")
    if pa and pa not in ("", "0 (installed after CVE published)"):
        try:
            age = int(pa)
            if age > 90:
                parts.append(f"Software has been installed {age} days — long-standing exposure.")
        except ValueError:
            pass
    if not parts:
        parts.append(f"Remediate per standard {sev} patch SLA. Verify with vendor advisory.")
    if pub:
        parts.append(f"Vendor: {pub}.")
    return " ".join(parts)
